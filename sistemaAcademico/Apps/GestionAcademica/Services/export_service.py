"""
Servicio de exportación de datos a Excel.
Proporciona funcionalidad para exportar datos del sistema a formato Excel con streaming.
"""
import logging
from typing import List, Dict, Any, Iterator
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExportService:
    """Servicio para exportar datos a Excel con formato profesional y streaming."""
    
    # Colores del tema
    HEADER_COLOR = "4472C4"  # Azul
    HEADER_FONT_COLOR = "FFFFFF"  # Blanco
    ALT_ROW_COLOR = "F2F2F2"  # Gris claro
    
    # Límite de registros para usar streaming
    STREAMING_THRESHOLD = 1000
    
    @staticmethod
    def _should_use_streaming(queryset) -> bool:
        """Determina si se debe usar streaming basado en el tamaño del queryset."""
        try:
            count = queryset.count()
            return count > ExportService.STREAMING_THRESHOLD
        except:
            return False
    
    @staticmethod
    def exportar_estudiantes(queryset, filtros: Dict[str, str] = None) -> HttpResponse:
        """
        Exporta estudiantes a Excel con streaming para grandes volúmenes.
        
        Args:
            queryset: QuerySet de estudiantes
            filtros: Diccionario con filtros aplicados
            
        Returns:
            HttpResponse con el archivo Excel
        """
        # Usar iterator() para evitar cargar todo en memoria
        # chunk_size controla cuántos registros se cargan a la vez
        CHUNK_SIZE = 500
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"
        
        # Información del reporte
        ws['A1'] = 'REPORTE DE ESTUDIANTES'
        ws['A1'].font = Font(size=16, bold=True, color=ExportService.HEADER_COLOR)
        ws['A2'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(size=10, italic=True)
        
        # Filtros aplicados
        if filtros:
            row = 3
            if filtros.get('search'):
                ws[f'A{row}'] = f'Búsqueda: {filtros["search"]}'
                ws[f'A{row}'].font = Font(size=10, italic=True)
                row += 1
        
        # Encabezados (fila 5)
        headers = ['ID', 'Nombres', 'Apellidos', 'Identificación', 'Estado']
        header_row = 5
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color=ExportService.HEADER_FONT_COLOR)
            cell.fill = PatternFill(start_color=ExportService.HEADER_COLOR, 
                                   end_color=ExportService.HEADER_COLOR, 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Datos - usar iterator() para streaming
        row_num = header_row + 1
        total_registros = 0
        
        # Procesar en chunks para evitar cargar todo en memoria
        for estudiante in queryset.iterator(chunk_size=CHUNK_SIZE):
            # Alternar color de filas
            fill_color = ExportService.ALT_ROW_COLOR if row_num % 2 == 0 else "FFFFFF"
            
            data = [
                estudiante.get('id_persona') if isinstance(estudiante, dict) else estudiante.id_persona,
                estudiante.get('nombres') if isinstance(estudiante, dict) else estudiante.nombres,
                estudiante.get('apellidos') if isinstance(estudiante, dict) else estudiante.apellidos,
                estudiante.get('identificacion') if isinstance(estudiante, dict) else estudiante.identificacion,
                'Activo'
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.fill = PatternFill(start_color=fill_color, 
                                       end_color=fill_color, 
                                       fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # Alineación
                if col_num == 1:  # ID centrado
                    cell.alignment = Alignment(horizontal="center")
            
            row_num += 1
            total_registros += 1
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 10,  # ID
            'B': 25,  # Nombres
            'C': 25,  # Apellidos
            'D': 15,  # Identificación
            'E': 12,  # Estado
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Agregar totales
        total_row = row_num + 1
        ws[f'A{total_row}'] = 'TOTAL DE REGISTROS:'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = total_registros
        ws[f'B{total_row}'].font = Font(bold=True)
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'estudiantes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        logger.info(f"Exportación de estudiantes completada: {total_registros} registros")
        
        return response
    
    @staticmethod
    def exportar_usuarios(queryset, filtros: Dict[str, str] = None) -> HttpResponse:
        """
        Exporta usuarios a Excel con streaming para grandes volúmenes.
        
        Args:
            queryset: QuerySet de usuarios
            filtros: Diccionario con filtros aplicados
            
        Returns:
            HttpResponse con el archivo Excel
        """
        # Usar iterator() para evitar cargar todo en memoria
        CHUNK_SIZE = 500
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        # Información del reporte
        ws['A1'] = 'REPORTE DE USUARIOS DEL SISTEMA'
        ws['A1'].font = Font(size=16, bold=True, color=ExportService.HEADER_COLOR)
        ws['A2'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(size=10, italic=True)
        
        # Filtros aplicados
        row = 3
        if filtros:
            if filtros.get('search'):
                ws[f'A{row}'] = f'Búsqueda: {filtros["search"]}'
                ws[f'A{row}'].font = Font(size=10, italic=True)
                row += 1
            if filtros.get('tipo_usuario'):
                ws[f'A{row}'] = f'Tipo de Usuario: {filtros["tipo_usuario"]}'
                ws[f'A{row}'].font = Font(size=10, italic=True)
                row += 1
        
        # Encabezados
        headers = ['ID', 'Usuario', 'Nombre Completo', 'Tipo de Usuario', 'Estado']
        header_row = row + 2
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color=ExportService.HEADER_FONT_COLOR)
            cell.fill = PatternFill(start_color=ExportService.HEADER_COLOR, 
                                   end_color=ExportService.HEADER_COLOR, 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Datos - usar iterator() para streaming
        row_num = header_row + 1
        total_registros = 0
        
        # Procesar en chunks para evitar cargar todo en memoria
        for usuario in queryset.iterator(chunk_size=CHUNK_SIZE):
            # Alternar color de filas
            fill_color = ExportService.ALT_ROW_COLOR if row_num % 2 == 0 else "FFFFFF"
            
            nombre_completo = f"{usuario.id_persona.nombres} {usuario.id_persona.apellidos}"
            
            data = [
                usuario.id_usuario,
                usuario.usuario,
                nombre_completo,
                str(usuario.id_genr_tipo_usuario),
                'Activo'
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.fill = PatternFill(start_color=fill_color, 
                                       end_color=fill_color, 
                                       fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # Alineación
                if col_num == 1:  # ID centrado
                    cell.alignment = Alignment(horizontal="center")
            
            row_num += 1
            total_registros += 1
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 10,  # ID
            'B': 20,  # Usuario
            'C': 35,  # Nombre Completo
            'D': 20,  # Tipo de Usuario
            'E': 12,  # Estado
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Agregar totales
        total_row = row_num + 1
        ws[f'A{total_row}'] = 'TOTAL DE REGISTROS:'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = total_registros
        ws[f'B{total_row}'].font = Font(bold=True)
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'usuarios_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        logger.info(f"Exportación de usuarios completada: {total_registros} registros")
        
        return response
    
    @staticmethod
    def exportar_empleados(queryset, filtros: Dict[str, str] = None) -> HttpResponse:
        """
        Exporta empleados a Excel con streaming para grandes volúmenes.
        
        Args:
            queryset: QuerySet de empleados
            filtros: Diccionario con filtros aplicados
            
        Returns:
            HttpResponse con el archivo Excel
        """
        # Usar iterator() para evitar cargar todo en memoria
        CHUNK_SIZE = 500
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empleados"
        
        # Información del reporte
        ws['A1'] = 'REPORTE DE EMPLEADOS'
        ws['A1'].font = Font(size=16, bold=True, color=ExportService.HEADER_COLOR)
        ws['A2'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(size=10, italic=True)
        
        # Encabezados
        headers = ['ID', 'Nombres', 'Apellidos', 'Identificación', 'Tipo', 'Estado']
        header_row = 5
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color=ExportService.HEADER_FONT_COLOR)
            cell.fill = PatternFill(start_color=ExportService.HEADER_COLOR, 
                                   end_color=ExportService.HEADER_COLOR, 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Datos - usar iterator() para streaming
        row_num = header_row + 1
        total_registros = 0
        
        # Procesar en chunks para evitar cargar todo en memoria
        for empleado in queryset.iterator(chunk_size=CHUNK_SIZE):
            fill_color = ExportService.ALT_ROW_COLOR if row_num % 2 == 0 else "FFFFFF"
            
            data = [
                empleado.id_persona,
                empleado.nombres,
                empleado.apellidos,
                empleado.identificacion,
                str(empleado.id_genr_tipo_usuario),
                'Activo'
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.fill = PatternFill(start_color=fill_color, 
                                       end_color=fill_color, 
                                       fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if col_num == 1:
                    cell.alignment = Alignment(horizontal="center")
            
            row_num += 1
            total_registros += 1
        
        # Ajustar ancho de columnas
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        
        # Totales
        total_row = row_num + 1
        ws[f'A{total_row}'] = 'TOTAL DE REGISTROS:'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = total_registros
        ws[f'B{total_row}'].font = Font(bold=True)
        
        # Respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'empleados_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        logger.info(f"Exportación de empleados completada: {total_registros} registros")
        
        return response
